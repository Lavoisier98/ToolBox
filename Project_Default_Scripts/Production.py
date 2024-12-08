import pyautogui as pg
import pyperclip as pc
import time as tm
from openpyxl import load_workbook

#IMPORTANTE: Caso esteja com arquivo aberto, dá erro de Permission Denied
def imgSearching(imageName, *args, **kwargs):
    #Optionals: conf, heights, widths, gScale
    conf = kwargs.get("conf",0.85)
    heights = kwargs.get("heights",0.5)
    widths = kwargs.get("widths",0.5)
    gScale = kwargs.get("gScale",True)
    img = None
    while img == None:
        try:
            img = pg.locateOnScreen(imgFolder+imageName,grayscale=gScale,confidence=conf)
        except pg.ImageNotFoundException:
            img = None
            tm.sleep(0.3)
    return {
        "top": img.top + heights * img.height,
        "left": img.left + widths * img.width,
        "height": img.height,
        "width": img.width,
        "img": img
    }
def inputAdding(scenarioNo):
    #Adding data from the line to simulator (4 tabs para chegar na txtbox)
    imgSearching(imageName="\SimulatorOpened.png",conf=0.9)
    tm.sleep(0.9)
    pg.press(["tab"], presses=26)
    tm.sleep(0.15)
    previousInput = None
    #Filling the fields
    for eachInput in inputsData:
        previousInput = eachInput
        pg.press("tab")
        tm.sleep(0.15)
        if ws[eachInput["column"]+str(scenarioNo)].value:
            #Prevenção para caso de pc.copy não funcionar
            if ws[previousInput["column"]+str(scenarioNo)].value != ws[eachInput["column"]+str(scenarioNo)].value and previousValue != None:
                verify = False
                while not verify:
                    try:
                        while pc.paste() == previousValue:
                            pc.copy(ws[eachInput["column"]+str(scenarioNo)].value)
                            tm.sleep(0.02)
                        verify = True
                    except pc.PyperclipWindowsException:
                        verify = False
                        tm.sleep(0.6)
            else:
                verify = False
                while not verify:
                    pc.copy(ws[eachInput["column"]+str(scenarioNo)].value)
                    try:
                        previousValue = pc.paste()
                        verify = True
                    except pc.PyperclipWindowsException:
                        verify = False
            pg.hotkey("ctrl","v")
    pg.press(["enter","enter"])


textoResposta = pc.paste()
parameters = [
    {
        "index":"Message",
        "end":"Internal Message",
        "column":"J"
    },
    {
        "index":"Internal Message",
        "end":"Sold To No.",
        "column":"K"
    },
    {
        "index":"Sold To No.",
        "end":"Ship To No.",
        "column":"Y"
    },
    {
        "index":"Ship To No.",
        "end":"Material",
        "column":"Z"
    },
    {
        "index":"Material",
        "end":"Material Type",
        "column":"X"
    },
    {
        "index":"Material Type",
        "end":"Sold To Customer Type",
        "column":"V"
    },
    {
        "index":"Parent Material",
        "end":"Sales Organization",
        "column":"W"
    },
    {
        "index":"Sales Organization",
        "end":"Forwarding Agent",
        "column":"AB"
    },
    {
        "index":"Forwarding Agent",
        "end":"Price Date",
        "column":"AA"
    },
    {
        "index":"Plant",
        "end":"Sales Unit",
        "column":"L"
    },
    {
        "index":"Sales Unit",
        "end":"Sales Organization (SOD)",
        "column":"M"
    },
    {
        "index":"Ship Condition",
        "end":"From Country",
        "column":"N"
    },
    {
        "index":"From Country",
        "end":"To Country",
        "column":"O"
    },
    {
        "index":"Inco Terms1",
        "end":"Inco Terms2",
        "column":"P"
    },
    {
        "index":"Inco Terms2",
        "end":"Min",
        "column":"Q"
    },
    {
        "index":"Min",
        "end":"Mult",
        "column":"R",
        "start": 300
    },
    {
        "index":"Mult",
        "end":"Max",
        "column":"S",
        "start": 300
    },
    {
        "index":"Max",
        "end":"Min/Mult/Max Unit",
        "column":"T",
        "start": 300
    },
    {
        "index":"Min/Mult/Max Unit",
        "end":"Payment Term",
        "column":"U"
    }
]
inputsData = [
    {
        "index":"Sold To Customer",
        "column":"D",
        "waveColumn":"AA"
    },
    {
        "index":"Ship To Customer",
        "column":"F",
        "waveColumn":"AF"
    },
    {
        "index":"Material",
        "column":"B",
        "waveColumn":"W"
    },
    {
        "index":"Sales Organization",
        "column":"I",
        "waveColumn":"AK"
    },
    {
        "index":"Freight Forwarding Agent ZU",
        "column":"H",
        "waveColumn":"AI"
    }
]
expecteds = {
    "Incoterm": {
        "wave":"BS",
        "output":"AC",
        "parameter": 9
    },
    "Shipping Condition": {
        "wave":"BO",
        "output":"AD",
        "parameter": 11
    }
}
messageCategory = {
    "column":"AB"
}
messageContent = [
    {
        "category":"SOD Mismatch",
        "content":"SOD MISMATCH"
    },
    {
        "category":"Incoterm Exception",
        "content":"INCOTERM EXCEPTION"
    },
    {
        "category":"SC Error",
        "content":"SHIPPING CONDITION EXCEPTION"
    },
    {
        "category":"Missing Route",
        "content":"Missing Route"
    },
    {
        "category":"Missing Price",
        "content":"No Price"
    },
    {
        "category":"Success",
        "content":"Success"
    },
    {
        "category":"Price or Availability",
        "content":"PRICE&AVAILABILITY"
    },
    {
        "category":"No SOD",
        "content":":NO SOD"
    },
    {
        "category":"No Customer",
        "content":":NO Customer"
    },
    {
        "category":"Availability",
        "content":"Availability"
    },
    {
        "category":"Availability",
        "content":"AVAILABILITY"
    },
    {
        "category":"No Material",
        "content":"NO Material"
    },
    {
        "category":"Other",
        "content":""
    }
]

#AQUI JAZ A MONTAGEM DO CONTEÚDO DA WAVE 4 (IMPORTANTE TER COPIADO A ÚLTIMA VERSÃO DO ARQUIVO)
#wbWavePath = r'C:\Users\NE20119\OneDrive - The Dow Chemical Company\Desktop\Base de Dados\CPM DCC_Order Validation_Wave 4 (version 2).xlsx'
wbPath = r'C:\Users\NE20119\OneDrive - The Dow Chemical Company\Desktop\Base de Dados\Simulator_Tester.xlsx'


#Loadear o Workbook destino das informações do simulador
#wb é o Workbook, ws é a Worksheet, tabsStreak é quantidade de abas simultâneas
#wbWavePath = r'C:\Users\NE20119\OneDrive - The Dow Chemical Company\Desktop\Base de Dados\CPM DCC_Order Validation_Wave 4 (version 2).xlsx'
urlSimulator = r"https://www.dow.com/en-us/my-account.html?#/single-sim"
imgFolder = r'C:\Users\NE20119\OneDrive - The Dow Chemical Company\Desktop\Base de Dados\VS Code\Python\Images'
tela = pg.size()
wb = load_workbook(filename= wbPath)
#wbWave = load_workbook(filename= wbWavePath)
ws = wb["Scenarios"]
wsWave = wb["Simulate"]
iterator = 3
tabsStreak = 9
#Inserção de células uma a uma
while ws["B"+str(iterator)].value:
    if not ws["B"+str(iterator+tabsStreak-1)].value:
        while not ws["B"+str(iterator+tabsStreak-1)].value:
            tabsStreak -= 1

    #Inicio de ciclo abrindo uma nova aba quando não há outras abertas
    tm.sleep(0.5)
    pg.hotkey("win","r")
    tm.sleep(0.3)
    pc.copy("msedge "+urlSimulator)
    pg.hotkey("ctrl","v")
    tm.sleep(0.2)
    pg.press("enter")
    tm.sleep(0.2)
    inputAdding(iterator)
    #Com uma aba aberta, abre-se todas as outras
    for addRows in range(1,tabsStreak):
        pg.hotkey("ctrl","t")
        pc.copy(urlSimulator)
        pg.hotkey("ctrl","v")
        tm.sleep(0.1)
        pg.press("enter")
        tm.sleep(0.1)
        inputAdding(iterator+addRows)
        tm.sleep(0.1)
    
    #Retorno à primeira aba, pois tem as informações da primeira linha coletada, para seguir a ordem do Excel
    for _ in range(1,tabsStreak):
        pg.hotkey("fn","ctrl","pgup")
        tm.sleep(0.1)

    currentTab = 1
    while currentTab <= tabsStreak:
        #waiting until the simulator's loading
        imgSearching(imageName="\SimulatorLoaded.png")
        tm.sleep(0.2)

        #copy the whole table to move it to clipboard area to become a python variable (textoResposta)
        imgFound = imgSearching(imageName="\SimulatorLoaded.png",heights= 0.265,widths= -0.066)
        pg.moveTo(x=imgFound["left"],y=imgFound["top"])
        pg.mouseDown(x=imgFound["left"],y=imgFound["top"],button="left")
        pg.moveTo(x=0.92*tela.width,y=imgFound["top"],duration=0.2)
        pg.moveTo(x=0.93*tela.width,y=imgFound["top"],duration=0.1)
        pg.mouseUp()
        tm.sleep(0.2)
        pg.hotkey("ctrl","c")
        textoResposta = pc.paste()
        print(textoResposta)
        #AQUI JAZ A COLAGEM DO CONTEÚDO COPIADO PELO PYAUTOGUEI NO EXCEL
        for eachParameter in parameters:
            print(eachParameter["index"]+": "+str(textoResposta.find(eachParameter["index"])))
            ws[eachParameter["column"]+str(iterator)].value = \
                textoResposta[ \
                    textoResposta.find(eachParameter["index"])+ \
                    len(eachParameter["index"])+1:textoResposta.find(eachParameter["end"])].strip()
            textoResposta = textoResposta[textoResposta.find(eachParameter["index"])+len(eachParameter["index"])+1:].strip()
        
        #Verificar IncoTermException e ShippingConditionError
        #iteratorWave = 3
        #rowFound = False
        #while not rowFound:
        #    rowFound = True
        #    for indvInput in inputsData:
        #        if rowFound:
        #            print(indvInput["waveColumn"]+str(iteratorWave))
        #            if wsWave[indvInput["waveColumn"]+str(iteratorWave)].value == None:
        #                if ws[indvInput["column"]+str(iterator)].value != None:
        #                    rowFound = False
        #            elif ws[indvInput["column"]+str(iterator)].value == None:
        #                if wsWave[indvInput["waveColumn"]+str(iteratorWave)].value != None:
        #                    rowFound = False
        #            else:
        #                
        #                if str(wsWave[indvInput["waveColumn"]+str(iteratorWave)].value) != str(ws[indvInput["column"]+str(iterator)].value):
        #                    print(str(wsWave[indvInput["waveColumn"]+str(iteratorWave)].value)+" =/= "+str(ws[indvInput["column"]+str(iterator)].value))
        #                    rowFound = False
        #    if not rowFound:
        #        iteratorWave += 1
        #        if iteratorWave == 1000:
        #            iteratorWave = 1048577

        #for indvExp in expecteds.keys():
        #    if wsWave[expecteds[indvExp]["wave"]+str(iteratorWave)].value != None:
        #        ws[expecteds[indvExp]["output"]+str(iterator)].value=wsWave[expecteds[indvExp]["wave"]+str(iteratorWave)].value
        #    #print(parameters[9]["column"])
        #    print(parameters[int(expecteds[indvExp]["parameter"])])
        #    if ws[parameters[int(expecteds[indvExp]["parameter"])]["column"]+str(iterator)] and wsWave[expecteds[indvExp]["wave"]+str(iteratorWave)]:
        #        if ws[parameters[expecteds[indvExp]["parameter"]]["column"]+str(iterator)] != wsWave[expecteds[indvExp]["wave"]+str(iteratorWave)]:
        #            ws[parameters[0]["column"]+str(iterator)].value = ws[parameters[0]["column"]+str(iterator)].value+":"+indvExp.upper()+" EXCEPTION"

        #Classificação da mensagem
        #for indvCategory in messageContent:
        #    if ws[messageCategory["column"]+str(iterator)].value != None:
        #        if indvCategory["content"] in ws[parameters[0]["column"]+str(iterator)].value+ws[parameters[1]["column"]+str(iterator)].value:
        #            ws[messageCategory["column"]+str(iterator)].value = indvCategory["content"]
        
        iterator += 1
        currentTab += 1
        pg.hotkey("ctrl","w")
        if not ws["B"+str(iterator)].value:
            break
    wb.save(filename= wbPath)
wb.save(filename= wbPath)
wb.close
print("Workbook salvo com sucesso")

#tm.sleep(2)
#pg.hotkey('alt','tab')
#tm.sleep(2)

#pg.hotkey('alt','tab')
#pc.copy("AAAA")
#print("Feito")