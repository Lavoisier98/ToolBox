import pyautogui as pg
import pyperclip as pc
import time as tm
from openpyxl import load_workbook

#IMPORTANTE: Caso esteja com arquivo aberto, dá erro de Permission Denied
def imgSearching(imageName, *args, **kwargs):
    #Optionals: conf, heights, widths, gScale
    conf = kwargs.get("conf",0.85)
    heights = kwargs.get("heights",0.5)
    widths = kwargs.get("widths",0.5)
    gScale = kwargs.get("gScale",True)
    img = None
    #print("esteve aqui 1b")
    while img == None:
        try:
            img = pg.locateOnScreen(imgFolder+imageName,grayscale=gScale,confidence=conf)
        except pg.ImageNotFoundException:
            img = None
            tm.sleep(0.3)
    #print("esteve aqui 1c")
    return {
        "top": img.top + heights * img.height,
        "left": img.left + widths * img.width,
        "height": img.height,
        "width": img.width,
        "img": img
    }
def inputAdding(scenarioNo):
    #Adding data from the line to simulator (4 tabs para chegar na txtbox)
    imgSearching(imageName="\BStructureOpen.png",conf=0.9)
    tm.sleep(0.2)
    pg.press(["tab", "tab", "tab", "tab", "tab", "tab", "tab", "tab", "tab", "enter", "down", "enter"])
    tm.sleep(0.15)
    previousInput = None
    #Filling the fields
    for eachInput in inputsData:
        pg.press("tab")
        tm.sleep(0.15)
        if ws[eachInput["column"]+str(scenarioNo)].value:
            if previousInput == None:
                previousInput = eachInput
            #Prevenção para caso de pc.copy não funcionar
            if ws[previousInput["column"]+str(scenarioNo)].value != ws[eachInput["column"]+str(scenarioNo)].value and previousValue != None:
                verify = False
                while not verify:
                    try:
                        while pc.paste() == previousValue:
                            pc.copy(ws[eachInput["column"]+str(scenarioNo)].value)
                            tm.sleep(0.02)
                        verify = True
                    except pc.PyperclipWindowsException:
                        verify = False
                        tm.sleep(0.6)
            else:
                verify = False
                while not verify:
                    pc.copy(ws[eachInput["column"]+str(scenarioNo)].value)
                    try:
                        previousValue = pc.paste()
                        verify = True
                    except pc.PyperclipWindowsException:
                        verify = False
                        tm.sleep(0.6)
            previousInput = eachInput
            pg.hotkey("ctrl","v")
    pg.press(["tab","enter","enter"])


textoResposta = pc.paste()
parameters = [
    {
        "index":"Market Facing Business Structure Search Results for GMID:",
        "end":"Management Group Code",
        "column":"B"
    },
    {
        "index":"Management Group Code",
        "end":"Management Group Name",
        "column":"C"
    },
    {
        "index":"Management Group Name",
        "end":"Business Group",
        "column":"D"
    },
    {
        "index":"Business Group",
        "end":"Business Group Name",
        "column":"E"
    },
    {
        "index":"Business Group Name",
        "end":"Business",
        "column":"F"
    },
    {
        "index":"Business",
        "end":"Business Name",
        "column":"G"
    },
    {
        "index":"Business Name",
        "end":"Value Center",
        "column":"H"
    },
    {
        "index":"Value Center",
        "end":"Value Center Name",
        "column":"I"
    },
    {
        "index":"Value Center Name",
        "end":"Performance Center",
        "column":"J"
    },
    {
        "index":"Profit Center / Profit Center Classification",
        "end":"Profit Center Name",
        "column":"K"
    },
    {
        "index":"Profit Center Name",
        "end":"Plan Product",
        "column":"L"
    },
    {
        "index":"Plan Product",
        "end":"Plan Product Name",
        "column":"M"
    },
    {
        "index":"Plan Product Name",
        "end":"Trade Product",
        "column":"N"
    },
    {
        "index":"Trade Product",
        "end":"Trade Product Description",
        "column":"O"
    },
    {
        "index":"Trade Product Description",
        "end":"Material",
        "column":"P"
    },
    {
        "index":"Material",
        "end":"Material Description",
        "column":"Q"
    },
    {
        "index":"Material Description",
        "end":"Material Type",
        "column":"R"
    },
    {
        "index":"Material Type",
        "end":"Authorization Group (QAC)",
        "column":"S"
    },
    {
        "index":"Authorization Group (QAC)",
        "end":"################",
        "column":"T"
    }
]
inputsData = [
    {
        "index":"GMID",
        "column":"A",
        "waveColumn":"W"
    }
]
#AQUI JAZ A MONTAGEM DO CONTEÚDO DA WAVE 4 (IMPORTANTE TER COPIADO A ÚLTIMA VERSÃO DO ARQUIVO)
#wbWavePath = r'C:\Users\NE20119\OneDrive - The Dow Chemical Company\Desktop\Base de Dados\CPM DCC_Order Validation_Wave 4 (version 2).xlsx'
wbPath = r'C:\Users\NE20119\OneDrive - The Dow Chemical Company\Desktop\Base de Dados\Simulator_Tester.xlsx'


#Loadear o Workbook destino das informações do simulador
#wb é o Workbook, ws é a Worksheet, tabsStreak é quantidade de abas simultâneas
urlSimulator = r"https://ppdcsc.dow.com/apps/business%20structure/Market_Facing_Search.asp"
imgFolder = r'C:\Users\NE20119\OneDrive - The Dow Chemical Company\Desktop\Base de Dados\VS Code\Python\Images'
tela = pg.size()
wb = load_workbook(filename= wbPath)
#wbWave = load_workbook(filename= wbWavePath)
ws = wb["Bstructure"]
wsWave = wb["Simulate"]
iterator = 3
tabsStreak = 9
#Inserção de células uma a uma
while ws["A"+str(iterator)].value:
    #Redefinição de tabStreak caso não haja quantidade inicial
    if not ws["A"+str(iterator+tabsStreak-1)].value:
        while not ws["A"+str(iterator+tabsStreak-1)].value:
            tabsStreak -= 1

    #Inicio de ciclo abrindo uma nova aba quando não há outras abertas
    tm.sleep(0.5)
    pg.hotkey("win","r")
    tm.sleep(0.3)
    pc.copy("msedge "+urlSimulator)
    pg.hotkey("ctrl","v")
    tm.sleep(0.2)
    pg.press("enter")
    tm.sleep(0.2)
    inputAdding(iterator)

    #Com uma aba aberta, abre-se todas as outras
    for addRows in range(1,tabsStreak):
        pg.hotkey("ctrl","t")
        pc.copy(urlSimulator)
        pg.hotkey("ctrl","v")
        tm.sleep(0.1)
        pg.press("enter")
        tm.sleep(0.1)
        inputAdding(iterator+addRows)
        tm.sleep(0.1)
    
    #Retorno à primeira aba, pois tem as informações da primeira linha coletada, para seguir a ordem do Excel
    for _ in range(1,tabsStreak):
        pg.hotkey("fn","ctrl","pgup")
        tm.sleep(0.1)

    currentTab = 1
    while currentTab <= tabsStreak:

        #copy the whole table to move it to clipboard area to become a python variable (textoResposta)
        #print("esteve aqui 1")
        imgFound = imgSearching(imageName="\BStructureOpen4.png",widths=0.007,heights=0.19)
        #print("esteve aqui 2")
        pg.moveTo(x=imgFound["left"],y=imgFound["top"])
        pg.mouseDown(x=imgFound["left"],y=imgFound["top"],button="left")
        pg.moveTo(x=0.92*tela.width,y=imgFound["top"],duration=0.2)
        pg.moveTo(x=0.94*tela.width,y=0.95*tela.height,duration=0.1)
        pg.mouseUp()
        tm.sleep(0.2)
        pg.hotkey("ctrl","c")
        textoResposta = pc.paste()
        print(textoResposta)

        #AQUI JAZ A COLAGEM DO CONTEÚDO COPIADO PELO PYAUTOGUEI NO EXCEL
        for eachParameter in parameters:
            ws[eachParameter["column"]+str(iterator)].value = \
                textoResposta[ \
                    textoResposta.find(eachParameter["index"])+ \
                    len(eachParameter["index"])+1: \
                    textoResposta.find( \
                        eachParameter["end"], \
                        textoResposta.find(eachParameter["index"])+ \
                        len(eachParameter["index"])+1 \
                    ) \
                ].strip()
            textoResposta = textoResposta[textoResposta.find(eachParameter["index"])+len(eachParameter["index"])+1:].strip()
        iterator += 1
        currentTab += 1
        tm.sleep(0.2)
        pg.hotkey("ctrl","w")
        tm.sleep(0.2)
        if not ws[inputsData[0]["column"]+str(iterator)].value:
            break
    wb.save(filename= wbPath)
wb.save(filename= wbPath)
wb.close
print("Workbook salvo com sucesso")
